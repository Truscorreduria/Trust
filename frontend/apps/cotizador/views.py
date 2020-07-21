from django.shortcuts import render, HttpResponse, render_to_response, Http404, HttpResponseRedirect
from django.http.response import JsonResponse
from backend.signals import *
from backend.utils import calcular_tabla_pagos
from grappelli_extras.utils import Codec
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from backend.forms import *
from django.template.loader import render_to_string
from easy_pdf.rendering import render_to_pdf_response, render_to_pdf
from datetime import datetime
from django.db.models import Count, Avg
from utils.utils import send_email
from django.core.exceptions import ValidationError
from django.contrib.auth.password_validation import validate_password
from django.conf import settings
import os
from django.db.models import Sum
from backend.decorators import profile_required
import random
import string
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
from django.contrib.auth import update_session_auth_hash
import secrets
from django.template import Template, Context
from django.contrib import messages


# region views cotizador

@profile_required
@login_required(login_url="/cotizador/login/")
def inicio(request):
    context = {}
    # context = {'productos': Producto.objects.filter(active=True),
    #            'marcas': marcas, 'modelos': modelos,
    #            'aseguradoras': Aseguradora.objects.filter(active=True),
    #            'producto': Producto.objects.get(code='0001'),
    #            'perfil_form': ProfileForm(instance=request.user),
    #            'today': datetime.now()
    #            }
    if request.method == "POST":
        form = ProfileForm(request.POST)
        context['perfil_form'] = form
        if form.is_valid():
            data = form.cleaned_data
            user = request.user
            user.first_name = data['first_name']
            user.last_name = data['last_name']
            user.email = data['email']
            user.save()
    return render(request, 'cotizador/base.html', context=context)


def _prepare_marca(marca):
    return (marca['marca'], marca['marca'])


def _prepare_modelo(modelo):
    return (modelo['modelo'], modelo['modelo'])


def _prepare_annos(anno):
    return (anno['anno'], anno['anno'])


@profile_required
@login_required(login_url="/cotizador/login/")
def cotizar(request):
    config = get_config(request.user)
    soa_descontado = round((config.soa_automovil * (1 - config.soa_descuento)), 2)
    today = datetime.now()
    year = int(str(today.year))
    marcas = Referencia.objects.values('marca').annotate(annos=Count('anno')).order_by('marca')
    annos = Referencia.objects.filter(marca=marcas[0]['marca']).values('anno').annotate(
        Count('valor')).order_by('marca', 'anno')
    modelos = Referencia.objects.filter(marca=marcas[0]['marca'], anno=annos[0]['anno']).values('modelo').annotate(
        Count('valor')).order_by('marca', 'anno', 'modelo')
    # annos = range(year - 10, year + 2)
    annos = [_prepare_annos(a) for a in annos]
    marcas = [_prepare_marca(marca) for marca in marcas]
    modelos = [_prepare_modelo(modelo) for modelo in modelos]
    context = {
        'marcas': marcas, 'modelos': modelos, 'annos': annos,
        'aseguradoras': Aseguradora.objects.filter(active=True),
        'anno_actual': year, 'soa_descontado': soa_descontado
    }
    return render(request, 'cotizador/cotizar.html', context)


def cotizacion_manual(request):
    nombres = request.user.first_name
    apellidos = request.user.last_name
    telefono = ""
    email = request.user.email
    marca = request.POST.get('marca')
    modelo = request.POST.get('modelo')
    anno = request.POST.get('anno')
    chasis = request.POST.get('chasis')
    motor = request.POST.get('motor')
    color = request.POST.get('color')
    placa = request.POST.get('placa')
    uso = request.POST.get('uso')
    html = render_to_string('cotizador/email/notificacion.html', {
        'nombres': nombres, 'apellidos': apellidos, 'telefono': telefono,
        'email': email, 'marca': marca, 'modelo': modelo, 'anno': anno,
        'chasis': chasis, 'motor': motor, 'color': color, 'placa': placa,
        'uso': uso
    })
    circulacion = request.FILES['file_circulacion']
    files = [("attachment", (circulacion.name, circulacion)), ]
    receipt = "cesarabel@deltacopiers.com,ventas@trustcorreduria.com,gcarrion@trustcorreduria.com"
    subject = "Se ha recibido una solicitud de cotización manual"
    send_email(subject=subject, receipt=receipt, html=html, files=files)
    ticket = Tramite()
    ticket.user = request.user
    ticket.descripcion = "COTIZACIÓN DE SEGURO DE AUTOMOVIL"
    ticket.marca = marca
    ticket.modelo = modelo
    ticket.anno = anno
    ticket.chasis = chasis
    ticket.motor = motor
    ticket.circulacion = circulacion
    ticket.placa = placa
    ticket.color = color
    ticket.save()
    return JsonResponse({})


@csrf_exempt
def get_annos(request):
    marca = request.POST.get('marca')
    annos = Referencia.objects.filter(marca=marca).values(
        'anno').annotate(models.Count('valor')).order_by('anno')
    data = [x['anno'] for x in annos]
    return JsonResponse(data, safe=False)


@csrf_exempt
def get_modelos(request):
    marca = request.POST.get('marca')
    anno = request.POST.get('anno')
    modelos = Referencia.objects.filter(marca=marca, anno=anno).values(
        'modelo').annotate(models.Count('valor')).order_by('modelo')
    data = [x['modelo'] for x in modelos]
    return JsonResponse(data, safe=False)


# @csrf_exempt
# def calcular_coberturas(coberturas, valor_nuevo, anno):
#     data = []
#     for c in coberturas:
#         for a in Aseguradora.objects.filter(active=True):
#             obj = {}
#             obj['cobertura'] = c.id
#             obj['aseguradora'] = a.id
#             p = c.get_precio(aseguradora=a)
#             if p.available:
#                 if c.tipo_exceso == '0.0':
#                     obj['valor'] = 0.0
#                 if c.tipo_exceso == 'valor_nuevo':
#                     obj['valor'] = calcular_exceso(valor_nuevo, c.tipo_calculo, p.valor)
#                 if c.tipo_exceso == 'valor_depreciado':
#                     valor_depreciado = a.depreciar(valor_nuevo, anno, 1)
#                     obj['valor'] = calcular_exceso(valor_depreciado, c.tipo_calculo, p.valor)
#                 data.append(obj)
#     return data


def el_mas_probable(referencias):
    average = referencias.aggregate(Avg('valor'))['valor__avg']
    elegido = referencias[0]
    diff = abs(elegido.valor - average)
    for r in referencias:
        ndiff = abs(r.valor - average)
        if ndiff < diff:
            diff = ndiff
            elegido = r
    return elegido


def anno_mas_cercano(years, year):
    diff = None
    selected = None
    for y in years:
        _diff = abs(int(y) - int(year))
        if diff:
            if _diff < diff:
                selected = y
                diff = _diff
        else:
            diff = _diff
            selected = y
    return selected


@csrf_exempt
def get_data(request):
    config = get_config(request.user)
    marca = request.POST.get('marca')
    porcentaje_deducible = config.porcentaje_deducible
    porcentaje_deducible_extension = config.porcentaje_deducible_extencion_territorial
    minimo_deducible = config.minimo_deducible
    minimo_deducible_extension = config.minimo_deducible
    porcentaje = str(round(config.porcentaje_deducible * 100, 0)).replace('.0', '%')
    porcentaje_extension = str(round(config.porcentaje_deducible_extencion_territorial * 100, 0)).replace('.0', '%')
    if marca in Marca.objects.all().values_list('marca', flat=True):
        marca_recargo = Marca.objects.get(marca=marca)
        minimo_deducible = marca_recargo.minimo
        porcentaje_deducible = marca_recargo.porcentaje_deducible
        porcentaje_deducible_extension = marca_recargo.porcentaje_deducible
        minimo_deducible_extension = marca_recargo.minimo
        porcentaje = marca_recargo.porcentaje()
        porcentaje_extension = marca_recargo.porcentaje()
        deducible_rotura_vidrios = marca_recargo.rotura_vidrios or 0.0
    else:
        deducible_rotura_vidrios = 0.0
    modelo = request.POST.get('modelo')
    anno = request.POST.get('anno')
    chasis = request.POST.get('chasis', None)
    exceso = float(request.POST.get('exceso'))
    chasis_encontrado = False

    try:
        if chasis and chasis != '':
            valor_nuevo = Referencia.objects.get(chasis=chasis)
            chasis_encontrado = True
        else:
            raise Exception()
    except:
        referencias = Referencia.objects.filter(marca=marca, modelo=modelo, anno=anno)
        if referencias.count() > 0:
            valor_nuevo = el_mas_probable(referencias)
        else:
            referencias = Referencia.objects.filter(marca=marca, modelo=modelo, anno__lt=anno).order_by('-anno')
            if referencias.count() > 0:
                _year = referencias[0].anno
                referencias = referencias.filter(anno=_year)
                valor_nuevo = el_mas_probable(referencias).to_json()
            else:
                referencias = Referencia.objects.filter(marca=marca, modelo=modelo).order_by('-anno')
                if referencias.count() > 0:
                    years = [x.anno for x in referencias.distinct('anno')]
                    _year = anno_mas_cercano(years, anno)
                    referencias = referencias.filter(anno=_year)
                    valor_nuevo = el_mas_probable(referencias).to_json()
                else:
                    valor_nuevo = {'marca': marca, 'modelo': modelo, 'anno': anno, 'valor': 0.0}

    aseguradora = config.aseguradora_automovil
    prima = round((valor_nuevo['valor'] * config.tasa_automovil) / 1000, 2)
    exceso = round((exceso / 2500) * 50, 2)
    vidrios = round((valor_nuevo['valor'] * 5) / 100, 2)
    prima_total = prima + config.soa_automovil + exceso
    emision = round(((prima + exceso) * aseguradora.emision) / 100, 2)
    iva = round(((prima + exceso + emision) * 15) / 100, 2)
    total = round(prima_total + emision + iva, 2)
    cuota = round(total / 12, 2)
    suma_asegurada = aseguradora.depreciar(valor_nuevo['valor'], anno)

    data = render_to_string('cotizador/txt/analitic_auto.html', {
        'referencia': valor_nuevo, 'aseguradora': aseguradora,
        'suma_asegurada': suma_asegurada, 'prima': prima, 'vidrios': vidrios,
        'prima_total': prima_total, 'emision': emision, 'iva': iva,
        'total': total, 'cuota': cuota, 'exceso': exceso
    })

    try:
        a, created = Analitics.objects.get_or_create(user=request.user, data=data,
                                                     area='cotizar auto', paso=2)
        a.save()
    except:
        pass

    return JsonResponse({'valor_nuevo': valor_nuevo.to_json(), 'suma_asegurada': suma_asegurada,
                         'prima': prima, 'vidrios': vidrios, 'prima_total': prima_total,
                         'emision': emision, 'iva': iva, 'total': total, 'cuota': cuota,
                         'exceso': exceso, 'chasis_encontrado': chasis_encontrado,
                         'porcentaje_deducible': porcentaje_deducible,
                         'porcentaje_deducible_extension': porcentaje_deducible_extension,
                         'minimo_deducible_extension': minimo_deducible_extension,
                         'minimo_deducible': minimo_deducible, 'porcentaje': porcentaje,
                         'porcentaje_extension': porcentaje_extension,
                         'deducible_rotura_vidrios': deducible_rotura_vidrios},
                        safe=False, encoder=Codec)


@csrf_exempt
def generar_cotizacion(request):
    user = request.user
    cliente = get_profile(user)
    ticket = Poliza()
    ticket.cliente = cliente
    ticket.fecha_emision = datetime.now()
    ticket.user = request.user
    ticket.nombres = request.POST.get('nombres', '')
    ticket.apellidos = request.POST.get('apellidos', '')
    ticket.email = request.POST.get('email', '')
    ticket.cedula = request.POST.get('cedula', '')
    ticket.telefono = request.POST.get('telefono', '')
    ticket.celular = request.POST.get('celular', '')
    ticket.domicilio = request.POST.get('domicilio', '')

    ticket.marca = request.POST.get('marca', '')
    ticket.modelo = request.POST.get('modelo', '')
    ticket.anno = request.POST.get('anno', '')
    ticket.chasis = request.POST.get('chasis', '')
    ticket.motor = request.POST.get('motor', '')
    ticket.placa = request.POST.get('placa', '')
    ticket.color = request.POST.get('color', '')
    ticket.valor_nuevo = request.POST.get('valor_nuevo', 0.0)
    ticket.suma_asegurada = request.POST.get('suma_asegurada', 0.0)
    ticket.subtotal = request.POST.get('prima_total', 0.0)
    ticket.emision = request.POST.get('emision', 0.0)
    ticket.iva = request.POST.get('iva', 0.0)
    ticket.total = request.POST.get('total_pagar', 0.0)
    ticket.porcentaje_deducible = request.POST.get('porcentaje_deducible')
    ticket.minimo_deducible = request.POST.get('minimo_deducible')
    ticket.porcentaje_deducible_extension = request.POST.get('porcentaje_deducible_extension')
    ticket.minimo_deducible_extension = request.POST.get('minimo_deducible_extension')
    ticket.deducible_rotura_vidrios = request.POST.get('deducible_rotura_vidrios')
    ticket.monto_exceso = float(request.POST.get('monto_exceso'))
    ticket.costo_exceso = float(request.POST.get('costo_exceso'))
    try:
        ticket.circulacion = request.FILES['circulacion']
    except:
        pass
    # ticket.save()

    context = {'poliza': ticket, 'config': get_config(user)}
    return render_to_pdf_response(request, 'cotizador/pdf/cotizacion.html', context)


@csrf_exempt
def guardar_poliza(request):
    config = get_config(request.user)
    cliente = get_profile(request.user)
    poliza = Poliza()
    poliza.procedencia = ProcedenciaPoliza.COTIZADOR
    str_date = request.POST.get('fecha_emision').split("-")
    now = datetime.now()
    fecha_emision = datetime(year=int(str_date[0]), month=int(str_date[1]), day=int(str_date[2]))
    if fecha_emision.year == now.year and fecha_emision.month == now.month \
            and fecha_emision.day == now.day:
        now = now + timedelta(hours=2)
        poliza.fecha_emision = now
    else:
        poliza.fecha_emision = fecha_emision

    cambio_asegurado = request.POST.get('cambio_asegurado')
    if cambio_asegurado == 'no':
        poliza.cliente = cliente
    else:
        c, created = ClienteNatural.objects.get_or_create(
            cedula=request.POST.get('cedula')
        )
        c.primer_nombre = request.POST.get('primer_nombre')
        c.segundo_nombre = request.POST.get('segundo_nombre')
        c.apellido_paterno = request.POST.get('apellido_paterno')
        c.apellido_materno = request.POST.get('apellido_materno')
        c.email_personal = request.POST.get('email')
        c.telefono = request.POST.get('telefono')
        c.celular = request.POST.get('celular')
        c.domicilio = request.POST.get('domicilio')
        c.save()
        poliza.cliente = c
    poliza.anno = request.POST.get('anno')
    poliza.marca = request.POST.get('marca')
    poliza.modelo = request.POST.get('modelo')
    poliza.user = request.user
    poliza.chasis = request.POST.get('chasis')
    poliza.motor = request.POST.get('motor')
    poliza.circulacion = request.POST.get('circulacion')
    poliza.placa = request.POST.get('placa')
    poliza.color = request.POST.get('color')
    poliza.uso = request.POST.get('uso')
    poliza.porcentaje_deducible = request.POST.get('porcentaje_deducible')
    poliza.minimo_deducible = request.POST.get('minimo_deducible')
    poliza.porcentaje_deducible_extension = request.POST.get('porcentaje_deducible_extension')
    poliza.minimo_deducible_extension = request.POST.get('minimo_deducible_extension')
    poliza.deducible_rotura_vidrios = request.POST.get('deducible_rotura_vidrios')
    poliza.monto_exceso = request.POST.get('monto_exceso')
    poliza.costo_exceso = request.POST.get('costo_exceso')

    # DEL CONFIG
    poliza.aseguradora = config.aseguradora_automovil
    poliza.ramo = config.ramo_automovil
    poliza.sub_ramo = config.sub_ramo_automovil
    poliza.contratante = config.empresa

    poliza.tipo_cobertura = request.POST.get('tipo_cobertura')
    if poliza.tipo_cobertura == '1':
        poliza.suma_asegurada = 0.00
        poliza.valor_nuevo = 0.00
        poliza.subtotal = round((config.soa_automovil * (1 - config.soa_descuento)), 2)
        poliza.emision = 0.00
        poliza.iva = 0.00
        poliza.total = round((config.soa_automovil * (1 - config.soa_descuento)), 2)
        poliza.monto_cuota = round((config.soa_automovil * (1 - config.soa_descuento)), 2)
    if poliza.tipo_cobertura == '2':
        poliza.suma_asegurada = float(request.POST.get('valor_depreciado'))
        poliza.valor_nuevo = float(request.POST.get('valor_nuevo'))
        poliza.subtotal = float(request.POST.get('prima_total'))
        poliza.emision = float(request.POST.get('emision'))
        poliza.iva = float(request.POST.get('iva'))
        poliza.total = float(request.POST.get('total_pagar'))

    poliza.medio_pago = request.POST.get('medio_pago')
    poliza.forma_pago = request.POST.get('forma_pago')
    if poliza.forma_pago == 'anual':
        poliza.cuotas = 1
        poliza.monto_cuota = float(request.POST.get('total_pagar'))
    if poliza.forma_pago == 'mensual':
        poliza.cuotas = int(request.POST.get('cuotas'))
        poliza.monto_cuota = round(poliza.total / poliza.cuotas, 2)

    if poliza.medio_pago == 'debito_automatico':
        poliza.moneda_cobro = request.POST.get('moneda_cobro')
        poliza.banco_emisor = request.POST.get('banco_emisor')
        poliza.card_number = request.POST.get('card_number')
        poliza.card_expiry = request.POST.get('card_expiry')
        poliza.card_type = request.POST.get('card_type')

    cesion = False
    if request.POST.get('cesion_derecho') == 'si':
        cesion = True
    poliza.cesion_derecho = cesion

    try:
        entidad = Entidad.objects.get(id=int(request.POST.get('entidad')))
    except:
        entidad = None
    poliza.beneficiario = entidad
    poliza.save()

    json_obj = dict()
    json_obj[config.fielmap_automovil('marca')] = request.POST.get('marca')
    json_obj[config.fielmap_automovil('modelo')] = request.POST.get('modelo')
    json_obj[config.fielmap_automovil('anno')] = request.POST.get('anno')
    json_obj[config.fielmap_automovil('chasis')] = request.POST.get('chasis')
    json_obj[config.fielmap_automovil('motor')] = request.POST.get('motor')
    json_obj[config.fielmap_automovil('placa')] = request.POST.get('placa')
    json_obj[config.fielmap_automovil('color')] = request.POST.get('color')

    dd = DatoPoliza()
    dd.poliza = poliza
    dd.extra_data = json.dumps(json_obj)
    dd.save()

    nueva_poliza.send(poliza, request=request)

    return JsonResponse(poliza.to_json(), encoder=Codec)


@csrf_exempt
def guardar_sepelio(request):
    config = get_config(request.user)
    o = OrdenTrabajo()
    o.tipo = "CF"
    o.user = request.user
    o.save()
    add_log(o, request=request)
    data = []
    for i, name in enumerate(request.POST.getlist('primer_nombre')):
        ben = benSepelio()
        ben.orden = o
        ben.empleado = request.user.profile()
        str_date = request.POST.getlist('fecha_nacimiento')[i].split("-")
        ben.fecha_nacimiento = datetime(year=int(str_date[0]), month=int(str_date[1]), day=int(str_date[2]))
        ben.parentesco = request.POST.getlist('parentesco')[i]
        ben.primer_nombre = request.POST.getlist('primer_nombre')[i]
        ben.segundo_nombre = request.POST.getlist('segundo_nombre')[i]
        ben.apellido_paterno = request.POST.getlist('apellido_paterno')[i]
        ben.apellido_materno = request.POST.getlist('apellido_materno')[i]
        ben.tipo_identificacion = request.POST.getlist('tipo_identificacion')[i]
        ben.cuotas = request.POST.getlist('cuotas')[i]
        ben.monto_cuota = request.POST.getlist('monto_cuota')[i]
        try:
            ben.file_cedula = request.FILES['documento_adjunto_%s' % i]
        except:
            pass
        ben.save()
        data.append(ben.to_json())

        poliza = Poliza.objects.get(cliente=config.empresa, estado_poliza=EstadoPoliza.ACTIVA,
                                    no_poliza=config.poliza_sepelio_dependiente)
        json_obj = dict()
        json_obj[config.fielmap_sepelio('primer_nombre')] = request.POST.getlist('primer_nombre')[i]
        json_obj[config.fielmap_sepelio('segundo_nombre')] = request.POST.getlist('segundo_nombre')[i]
        json_obj[config.fielmap_sepelio('apellido_paterno')] = request.POST.getlist('apellido_paterno')[i]
        json_obj[config.fielmap_sepelio('apellido_materno')] = request.POST.getlist('apellido_materno')[i]
        json_obj[config.fielmap_sepelio('fecha_nacimiento')] = request.POST.getlist('fecha_nacimiento')[i]
        json_obj[config.fielmap_sepelio('parentesco')] = request.POST.getlist('parentesco')[i]
        json_obj[config.fielmap_sepelio('costo')] = config.costo_sepelio
        json_obj[config.fielmap_sepelio('suma_asegurada')] = config.suma_sepelio

        dd = DatoPoliza()
        dd.poliza = poliza
        dd.extra_data = json.dumps(json_obj)
        dd.save()
    html = render_to_string('cotizador/email/notificacion_sepelio.html',
                            context={'object': o, 'opts': o._meta},
                            request=request)
    files = [("attachment", ("%s.%s" % (x.full_name(), get_extension(x.file_cedula)),
                             x.file_cedula.read())) for x in o.beneficiarios()]
    beneficiarios = o.beneficiarios()
    ot = render_to_pdf('cotizador/pdf/orden_trabajo_sepelio.html', {
        'beneficiarios': beneficiarios, 'now': o.created, 'config': config,
        'total_suma_asegurada': beneficiarios.aggregate(Sum('suma_asegurada'))['suma_asegurada__sum'],
        'total_costo': beneficiarios.aggregate(Sum('costo'))['costo__sum'], 'orden': o
    })

    files.append(("attachment", (o.nomeclatura() + ".pdf", ot)))

    send_email('Orden de Trabajo # %s' % o.nomeclatura(),
               config.email_sepelio, html=html, files=files)
    return JsonResponse({'beneficiarios': data, 'orden': o.to_json()}, encoder=Codec, safe=False)


@csrf_exempt
def costo_accidente(request):
    config = get_config(request.user)
    inicio = datetime(day=1, month=1, year=datetime.now().year)
    vence = datetime(year=inicio.year + 1, month=1, day=1)
    dias = (vence - datetime.now()).days
    prima = ((dias * config.costo_accidente) / (vence - inicio).days)
    costo = prima + config.costo_carnet_accidente
    emision = (costo) * 0.02
    return JsonResponse({'costo': round(costo, 2), 'emision': round(emision, 2),
                         'dias': dias, 'prima': round(prima, 2), 'carnet': config.costo_carnet_accidente,
                         'suma': config.suma_accidente_dependiente},
                        encoder=Codec)


@csrf_exempt
def guardar_accidente(request):
    config = get_config(request.user)
    o = OrdenTrabajo()
    o.tipo = "AP"
    o.user = request.user
    o.save()
    add_log(o, request=request)
    data = []
    for i, name in enumerate(request.POST.getlist('primer_nombre')):
        ben = benAccidente()
        ben.orden = o
        ben.empleado = request.user.profile()
        str_date = request.POST.getlist('fecha_nacimiento')[i].split("-")
        ben.fecha_nacimiento = datetime(year=int(str_date[0]), month=int(str_date[1]), day=int(str_date[2]))
        ben.parentesco = request.POST.getlist('parentesco')[i]
        ben.primer_nombre = request.POST.getlist('primer_nombre')[i]
        ben.segundo_nombre = request.POST.getlist('segundo_nombre')[i]
        ben.apellido_paterno = request.POST.getlist('apellido_paterno')[i]
        ben.apellido_materno = request.POST.getlist('apellido_materno')[i]
        ben.tipo_identificacion = request.POST.getlist('tipo_identificacion')[i]
        ben.suma_asegurada = request.POST.getlist('suma_asegurada')[i]
        ben.prima = request.POST.getlist('prima')[i]
        ben.carnet = request.POST.getlist('carnet')[i]
        ben.emision = request.POST.getlist('emision')[i]
        ben.cuotas = request.POST.getlist('cuotas')[i]
        ben.monto_cuota = request.POST.getlist('monto_cuota')[i]
        ben.costo = request.POST.getlist('costo')[i]
        try:
            ben.file_cedula = request.FILES['documento_adjunto_%s' % i]
        except:
            pass
        ben.save()
        data.append(ben.to_json())
        poliza = Poliza.objects.get(cliente=config.empresa, estado_poliza=EstadoPoliza.ACTIVA,
                                    no_poliza=config.poliza_accidente)
        json_obj = dict()
        json_obj[config.fielmap_accidente('primer_nombre')] = request.POST.getlist('primer_nombre')[i]
        json_obj[config.fielmap_accidente('segundo_nombre')] = request.POST.getlist('segundo_nombre')[i]
        json_obj[config.fielmap_accidente('apellido_paterno')] = request.POST.getlist('apellido_paterno')[i]
        json_obj[config.fielmap_accidente('apellido_materno')] = request.POST.getlist('apellido_materno')[i]
        json_obj[config.fielmap_accidente('fecha_nacimiento')] = request.POST.getlist('fecha_nacimiento')[i]
        json_obj[config.fielmap_accidente('parentesco')] = request.POST.getlist('parentesco')[i]
        json_obj[config.fielmap_sepelio('costo')] = config.costo_accidente
        json_obj[config.fielmap_sepelio('suma_asegurada')] = config.suma_accidente_dependiente

        dd = DatoPoliza()
        dd.poliza = poliza
        dd.extra_data = json.dumps(json_obj)
        dd.save()

    html = render_to_string('cotizador/email/notificacion_accidente.html',
                            context={'object': o, 'opts': o._meta},
                            request=request)
    files = [("attachment", ("%s.%s" % (x.full_name(), get_extension(x.file_cedula)),
                             x.file_cedula.read())) for x in o.beneficiarios()]
    beneficiarios = o.beneficiarios()
    ot = render_to_pdf('cotizador/pdf/orden_trabajo_accidente.html', {
        'beneficiarios': beneficiarios, 'total_prima': beneficiarios.aggregate(Sum('prima'))['prima__sum'],
        'total_suma_asegurada': beneficiarios.aggregate(Sum('suma_asegurada'))['suma_asegurada__sum'],
        'total_emision': beneficiarios.aggregate(Sum('emision'))['emision__sum'],
        'total_costo': beneficiarios.aggregate(Sum('costo'))['costo__sum'],
        'user': request.user, 'now': datetime.now(), 'orden': o
    })

    files.append(("attachment", (o.nomeclatura() + ".pdf", ot)))

    send_email('Orden de Trabajo # %s' % o.nomeclatura(),
               config.email_trust + config.email_accidente, html=html, files=files)
    return JsonResponse({'beneficiarios': data, 'orden': o.to_json()}, encoder=Codec, safe=False)


@csrf_exempt
def print_recibo(request):
    poliza = Poliza.objects.get(id=request.POST.get('id'))
    return render_to_pdf_response(request, 'cotizador/pdf/recibo.html', {
        'poliza': poliza
    })


@csrf_exempt
def print_cotizacion(request):
    poliza = Poliza.objects.get(id=request.POST.get('id'))
    return render_to_pdf_response(request, 'cotizador/pdf/cotizacion.html', {
        'poliza': poliza, 'fecha': datetime.now()
    })


@csrf_exempt
def print_condiciones(request):
    config = get_config(request.user)
    poliza = Poliza.objects.get(id=request.POST.get('id'))
    if poliza.tipo_cobertura == 'amplia':
        return render_to_pdf_response(request, 'cotizador/pdf/condiciones_particulares.html', {
            'poliza': poliza, 'config': config
        })
    if poliza.tipo_cobertura == 'basica':
        return render_to_pdf_response(request, 'cotizador/pdf/soa.html', {
            'poliza': poliza, 'config': config
        })


@csrf_exempt
def print_orden_trabajo(request):
    poliza = Poliza.objects.get(id=request.POST.get('id'))
    config = get_config(poliza.user)
    return render_to_pdf_response(request, 'cotizador/pdf/orden_trabajo.html', {
        'poliza': poliza, 'soa_descontado': round((config.soa_automovil * (1 - config.soa_descuento)), 2),
        'config': config
    })


@csrf_exempt
def print_documentos(request):
    poliza = Poliza.objects.get(id=request.POST.get('id'))
    return render_to_pdf_response(request, 'cotizador/pdf/documentos_auto.html', {
        'poliza': poliza
    })


@csrf_exempt
def print_orden_trabajo_sepelio(request):
    config = get_config(request.user)
    orden = OrdenTrabajo.objects.get(id=int(request.POST.get('orden')))
    beneficiarios = orden.beneficiarios()
    return render_to_pdf_response(request, 'cotizador/pdf/orden_trabajo_sepelio.html', {
        'beneficiarios': beneficiarios, 'now': datetime.now(), 'config': config,
        'total_suma_asegurada': beneficiarios.aggregate(Sum('suma_asegurada'))['suma_asegurada__sum'],
        'total_costo': beneficiarios.aggregate(Sum('costo'))['costo__sum'], 'orden': orden
    })


@csrf_exempt
def print_orden_trabajo_accidente(request):
    config = get_config(request.user)
    orden = OrdenTrabajo.objects.get(id=int(request.POST.get('orden')))
    beneficiarios = orden.beneficiarios()
    return render_to_pdf_response(request, 'cotizador/pdf/orden_trabajo_accidente.html', {
        'beneficiarios': beneficiarios, 'total_prima': beneficiarios.aggregate(Sum('prima'))['prima__sum'],
        'total_suma_asegurada': beneficiarios.aggregate(Sum('suma_asegurada'))['suma_asegurada__sum'],
        'total_emision': beneficiarios.aggregate(Sum('emision'))['emision__sum'],
        'total_costo': beneficiarios.aggregate(Sum('costo'))['costo__sum'],
        'user': request.user, 'now': datetime.now(), 'orden': orden, 'config': config
    })


@csrf_exempt
def print_soa(request):
    poliza = Poliza.objects.get(id=request.POST.get('id'))
    # poliza = Poliza.objects.get(id=95)

    return render_to_pdf_response(request, 'cotizador/pdf/soa.html', {
        'poliza': poliza
    })
    # return render(request, 'cotizador/pdf/soa.html', {'poliza': 'poliza'})


@csrf_exempt
def print_cesion(request):
    poliza = Poliza.objects.get(id=request.POST.get('id'))
    if poliza.cesion_derecho == True:
        return render_to_pdf_response(request, 'cotizador/pdf/cesion_derechos.html', {
            'poliza': poliza
        })
    else:
        return HttpResponse(status=204)


@csrf_exempt
def generar_solicitud(request):
    data = dict()
    data['aseguradora'] = request.POST.get('aseguradora')
    data['nombres'] = request.POST.get('nombres')
    data['apellidos'] = request.POST.get('apellidos')
    data['cedula'] = request.POST.get('cedula')
    data['telefono'] = request.POST.get('telefono')
    data['email'] = request.POST.get('email')
    data['anno'] = request.POST.get('anno')
    data['marca'] = request.POST.get('marca')
    data['modelo'] = request.POST.get('modelo')
    data['chasis'] = request.POST.get('chasis')
    data['chasis'] = "aqui"
    data['motor'] = request.POST.get('motor')
    data['circulacion'] = request.POST.get('circulacion')
    data['vin'] = request.POST.get('vin')
    data['certificado'] = request.POST.get('certificado')
    data['placa'] = request.POST.get('placa')
    data['color'] = request.POST.get('color')
    data['capacidad'] = request.POST.get('capacidad')
    data['tonelaje'] = request.POST.get('tonelaje')
    data['uso'] = request.POST.get('uso')
    data['valor_nuevo'] = request.POST.get('valor_nuevo')
    data['valor_depreciado'] = request.POST.get('valor_depreciado')
    return render_to_response('cotizador/pdf/solicitud.html', data)


@login_required(login_url="/cotizador/login/")
def perfil(request):
    form = ProfileForm(instance=request.user)
    context = {'perfil_form': form}
    if request.method == "POST":
        form = ProfileForm(request.POST)
        if form.is_valid():
            data = form.cleaned_data
            user = request.user
            user.email = data['email']
            perfil = user.profile()
            perfil.primer_nombre = data['primer_nombre']
            perfil.segundo_nombre = data['segundo_nombre']
            perfil.apellido_paterno = data['apellido_paterno']
            perfil.apellido_materno = data['apellido_materno']
            perfil.email_personal = data['email_personal']
            perfil.departamento_id = data['departamento']
            perfil.municipio_id = data['municipio']
            perfil.cedula = data['cedula']
            perfil.celular = data['celular']
            perfil.telefono = data['telefono']
            perfil.domicilio = data['domicilio']
            perfil.sucursal = data['sucursal']
            perfil.codigo_empleado = data['codigo_empleado']
            perfil.cargo = data['cargo']
            try:
                perfil.foto = request.FILES['foto']
            except:
                pass
            user.save()
            perfil.save()
            context['message'] = "Perfil actuliazado!"
    return render(request, 'cotizador/perfil.html', context=context)


@login_required
def change_password(request):
    user = request.user
    new_pass = request.POST.get('pass_new')
    profile = user.profile()
    try:
        validate_password(new_pass, user=user)
        if profile.cambiar_pass:
            if new_pass == request.POST.get('pass_conf'):
                user.set_password(new_pass)
                user.save()
                update_session_auth_hash(request, user)
                profile.cambiar_pass = False
                profile.save()
                return JsonResponse({'messages': ["Tu password se ha cambiado con exito!", ],
                                     'class': "success"})
            else:
                return JsonResponse({'messages': ["Los campos no coiciden!", ],
                                     'class': "danger"})
        else:
            if user.check_password(request.POST.get('pass_actual')):
                if new_pass == request.POST.get('pass_conf'):
                    user.set_password(new_pass)
                    user.save()
                    update_session_auth_hash(request, user)
                    return JsonResponse({'messages': ["Tu password se ha cambiado con exito!", ],
                                         'class': "success"})
                else:
                    return JsonResponse({'messages': ["Los campos no coiciden!", ],
                                         'class': "danger"})
            else:
                return JsonResponse({'messages': ["La contraseña actual es incorrecta!", ],
                                     'class': "danger"})
    except ValidationError as messages:
        return JsonResponse({'messages': list(messages),
                             'class': "danger"})


@profile_required
@login_required(login_url="/cotizador/login/")
def referenciados(request):
    return render(request, 'cotizador/referenciados.html')


@profile_required
@login_required(login_url="/cotizador/login/")
def misseguros(request):
    config = get_config(request.user)
    return render(request, 'cotizador/misseguros.html',
                  {'config': config})


@login_required(login_url="/cotizador/login/")
def solicitar_baja(request):
    cliente = get_profile(request.user)
    tipo = request.POST.get('tipo')
    dependiente = None
    referente = None
    if tipo == 'bensepelio':
        dependiente = benSepelio.objects.get(id=int(request.POST.get('beneficiario')))
        referente = "sepelio"
    if tipo == 'benaccidente':
        dependiente = benAccidente.objects.get(id=int(request.POST.get('beneficiario')))
        referente = "accidente"
    t = Tramite()
    t.user = request.user
    t.cliente = cliente
    t.descripcion = "Solicitud de baja de dependiente"
    t.nombres = " ".join([dependiente.empleado.primer_nombre, dependiente.empleado.segundo_nombre])
    t.apellidos = " ".join([dependiente.empleado.apellido_materno, dependiente.apellido_paterno])
    t.email = dependiente.empleado.email_personal
    t.cedula = dependiente.empleado.cedula
    t.telefono = dependiente.empleado.telefono
    t.domicilio = dependiente.empleado.domicilio
    t.referente = referente
    t.movimiento = '2'
    t.descripcion = "Exclusión del siguiente dependiente: %s" % dependiente.full_name()
    t.save()
    dependiente.ticket = t
    dependiente.save()

    NotificarTramite.send(t)

    return JsonResponse({'ticket': t.to_json()}, encoder=Codec)


@login_required(login_url="/cotizador/login/")
def solicitar_baja_auto(request):
    cliente = get_profile(request.user)
    poliza = Poliza.objects.get(id=int(request.POST.get('poliza')))
    t = Tramite()
    t.user = request.user
    t.cliente = poliza.cliente
    t.descripcion = "Solicitud de baja de seguro de vehículo"
    t.nombres = poliza.nombres
    t.apellidos = poliza.apellidos
    t.email = poliza.user.profile().email_personal
    t.cedula = poliza.user.profile().cedula
    t.telefono = poliza.user.profile().telefono
    t.domicilio = poliza.user.profile().domicilio
    t.referente = "auto"
    t.movimiento = '2'
    t.descripcion = "Solicutud de baja de seguro de vehículo"
    t.save()
    poliza.ticket = t
    poliza.save()

    NotificarTramite.send(t)
    return JsonResponse({'ticket': t.to_json()}, encoder=Codec)


@profile_required
@login_required(login_url="/cotizador/login/")
def contactanos(request):
    cliente = get_profile(request.user)
    if request.method == "POST":
        t = Tramite()
        t.user = request.user
        t.cliente = cliente
        t.nombres = request.POST.get('nombres', '')
        t.apellidos = request.POST.get('apellidos', '')
        t.email = request.POST.get('email', '')
        t.referente = request.POST.get('referente', '')
        t.movimiento = request.POST.get('movimiento', '')
        t.motivo = request.POST.get('motivo', '')
        t.descripcion = request.POST.get('comentarios', '')
        try:
            t.poliza = Poliza.objects.get(no_poliza=request.POST.get('no_poliza'))
        except:
            pass
        t.save()
        NotificarTramite.send(t)
    response = render(request, 'cotizador/contactanos.html')
    response.set_cookie('user', request.user.id)
    return response


def randomString(stringLength=10):
    """Generate a random string of fixed length """
    letters = string.ascii_lowercase
    return ''.join(random.choice(letters) for i in range(stringLength))


def restablecer_password(request):
    email = request.POST.get('email')

    try:
        user = User.objects.get(email=email)
    except:
        user = None

    if user:
        password = randomString()
        user.set_password(password)
        user.save()
        profile = user.profile()
        profile.cambiar_pass = True
        profile.save()
        html = render_to_string('cotizador/email/password_reset.html', {
            'password': password
        })
        send_email("Hemos restablecido tu contraseña", email, html)
        result = "success"
        message = "Contraseña restablecidad con éxito!"
    else:
        result = "error"
        message = "Usuario no encontrado."
    return JsonResponse({'result': result,
                         'message': message}, encoder=Codec)


def download(request):
    file_name = request.POST.get('file_name', request.GET.get('file_name'))
    path_to_file = os.path.join(settings.STATIC_ROOT, 'descargas', file_name)
    if os.path.exists(path_to_file):
        with open(path_to_file, 'rb') as fh:
            response = HttpResponse(fh.read(), content_type='application/force-download')
            response['Content-Disposition'] = 'inline; filename=%s' % os.path.basename(file_name)
            response['X-Sendfile'] = path_to_file
            return response
    else:
        raise Http404


def calcular_tabla_pagos_polizas(request):
    total = float(request.POST.get('total'))
    fecha_pago = datetime.strptime(request.POST.get('fecha'), '%d/%m/%Y')
    cuotas = int(request.POST.get('cuotas'))
    try:
        poliza = Poliza.objects.get(id=int(request.POST.get('poliza')))
    except:
        poliza = None
    monto_cuota = round(total / cuotas, 2)
    data = calcular_tabla_pagos(total, fecha_pago, cuotas, poliza)
    return JsonResponse(data, safe=False, encoder=Codec)


def reporte_inclusion(modelAdmin, request, queryset):
    wb = Workbook()
    ws = wb.active

    c = ws['A2']
    print(c)
    c.font = Font(size=11, bold=True, family=1)
    c.alignment = Alignment(horizontal='left')
    ws['A2'] = "Información necesaria para inclusión de deducciones."
    # ws.merge_cells('A2:G2')

    for cell in ws['A5:G5'][0]:
        cell.font = Font(underline="single", bold=True)
        cell.border = Border(left=Side(border_style='medium', color='000000'),
                             right=Side(border_style='medium', color='000000'),
                             top=Side(border_style='medium', color='000000'),
                             bottom=Side(border_style='medium', color='000000')
                             )
        cell.alignment = Alignment(horizontal='justify',
                                   vertical='bottom',
                                   text_rotation=0,
                                   wrap_text=True,
                                   shrink_to_fit=False,
                                   indent=0)
    ws['A5'] = "# Empl."
    ws['B5'] = "Nombre"
    ws['C5'] = "Monto"
    ws['D5'] = "Moneda"
    ws['E5'] = "Cant. Cuotas a deducir"
    ws['F5'] = "# POLIZA"
    ws['G5'] = "VIGENCIA"

    row = 6

    for obj in queryset:
        for dep in obj.beneficiarios():
            ws['A' + str(row)] = obj.user.profile().codigo_empleado
            ws['B' + str(row)] = dep.full_name()
            ws['C' + str(row)] = dep.costo
            ws['D' + str(row)] = "Dólares"
            ws['E' + str(row)] = ""
            ws['F' + str(row)] = dep.numero_poliza
            ws['G' + str(row)] = ""
            row += 1

    response = HttpResponse(content_type="application/ms-excel")
    content_disposition = "attachment; filename=Inclusiones.xlsx"
    response["Content-Disposition"] = content_disposition
    wb.save(response)
    return response


def ingresar_numero_poliza(request):
    p = Poliza.objects.get(id=int(request.POST.get('original')))
    p.no_poliza = request.POST.get('no_poliza')
    p.no_recibo = request.POST.get('no_recibo')
    p.save()
    poliza_lista.send(p, request=request)
    return JsonResponse(p.to_json(), encoder=Codec)


def enviar_contacto(request):
    config = get_config(request.user)
    ticket = Tramite.objects.get(id=int(request.POST.get('ticket')))
    html = render_to_string('cotizador/email/contacto_directo.html', {
        'ticket': ticket, 'comentarios': request.POST.get('comentarios')
    })
    send_email('Contacto directo', config.email_trust, html=html)
    return JsonResponse(ticket.to_json(), encoder=Codec)


def javascript(request, file):
    return render(request, 'cotizador/js/' + file)


def autorenovar_polizas():
    hoy = datetime.now()
    aseguradora = Aseguradora.objects.get(name='ASSA')
    for p in Poliza.objects.filter(valor_nuevo__gt=0):
        vence = datetime(year=p.fecha_vencimiento().year, month=p.fecha_vencimiento().month,
                         day=p.fecha_vencimiento().day)
        if vence <= hoy:
            p.suma_asegurada = aseguradora.depreciar(p.valor_nuevo, p.anno)
            emision = datetime(year=p.fecha_emision.year, month=p.fecha_emision.month, day=p.fecha_emision.day)
            p.fecha_emision = emision + timedelta(days=365)
            p.fecha_vence = emision + timedelta(days=730)
            p.save()


# endregion


# renovacion

def solicitud_renovacion_auto(request):
    poliza = Poliza.objects.get(id=int(request.POST.get('poliza')))
    s, _ = SolicitudRenovacion.objects.get_or_create(poliza=poliza)
    s.medio_pago = request.POST.get('medio-pago')
    s.forma_pago = request.POST.get('forma-pago')
    s.cuotas = request.POST.get('cantidad_cuotas')
    s.monto_cuota = request.POST.get('valor_cuota')
    s.save()
    return JsonResponse({})

    # 'medio-pago': ['deduccion_nomina'], 'forma-pago': ['mensual'], 'cantidad_cuotas': ['6'], 'valor_cuota': ['46.17']


def iniciar_proc():
    ps = Poliza.objects.filter(procedencia=ProcedenciaPoliza.COTIZADOR, fecha_emision__isnull=False,
                               cliente__isnull=False, fecha_vence__lte=datetime.now(), aseguradora__isnull=False,
                               estado_poliza=EstadoPoliza.ACTIVA)
    for p in ps:
        try:
            s = SolicitudRenovacion.objects.get(poliza=p)
        except:
            s = None
        nueva = RenovarPoliza.send(p, fecha_renovacion=p.fecha_vence)[0][1]
        nueva.user = p.user
        nueva.suma_asegurada = p.aseguradora.depreciar(p.valor_nuevo, p.anno)
        nueva.marca = p.marca
        nueva.modelo = p.modelo
        nueva.chasis = p.chasis
        nueva.anno = p.anno
        nueva.placa = p.placa
        nueva.color = p.color
        nueva.circulacion = p.circulacion
        nueva.tipo_cobertura = p.tipo_cobertura
        nueva.porcentaje_deducible = p.porcentaje_deducible
        nueva.porcentaje_deducible_extension = p.porcentaje_deducible_extension
        nueva.minimo_deducible = p.minimo_deducible
        nueva.minimo_deducible_extension = p.minimo_deducible_extension
        nueva.moneda = p.moneda
        nueva.costo_exceso = p.costo_exceso
        nueva.monto_exceso = p.monto_exceso
        nueva.valor_nuevo = p.valor_nuevo
        nueva.subtotal = p.subtotal
        nueva.descuento = p.descuento
        nueva.emision = p.emision
        nueva.iva = p.iva
        nueva.otros = p.otros
        nueva.total = p.total
        nueva.per_comision = p.per_comision
        nueva.amount_comision = p.amount_comision
        nueva.cesion_derecho = p.cesion_derecho
        nueva.beneficiario = p.beneficiario
        nueva.cesioinario = p.cesioinario
        nueva.forma_pago = p.forma_pago
        nueva.f_pago = p.f_pago
        nueva.medio_pago = p.medio_pago
        nueva.m_pago = p.m_pago
        nueva.cuotas = p.cuotas
        nueva.moneda_cobro = p.moneda_cobro
        nueva.banco_emisor = p.banco_emisor
        nueva.file_circulacion = p.file_circulacion
        nueva.estado_poliza = EstadoPoliza.PENDIENTE
        nueva.save()
        print(nueva)
        p.estado_poliza = EstadoPoliza.RENOVADA

        if s:
            nueva.forma_pago = s.forma_pago
            nueva.f_pago = s.f_pago
            nueva.medio_pago = s.medio_pago
            nueva.m_pago = s.m_pago
            nueva.cuotas = s.cuotas
        p.save()


def iniciar_notificacion_vencer():
    day = datetime.now() + timedelta(days=30)
    ps = Poliza.objects.filter(procedencia=ProcedenciaPoliza.COTIZADOR, fecha_emision__isnull=False,
                               cliente__isnull=False,
                               fecha_vence__year=day.year,
                               fecha_vence__month=day.month,
                               fecha_vence__day=day.day,
                               aseguradora__isnull=False,
                               estado_poliza=EstadoPoliza.ACTIVA)
    # ps = Poliza.objects.filter(fecha_vence__lte=day)
    for p in ps:
        print(p.no_poliza)
        config = p.get_config()
        if config:
            html = render_to_string('cotizador/email/notificacion_vence.html', {
                'body': config.email_texto, 'poliza': p
            })
            send_email('Tu póliza # %s está cerca de vencer' % p.no_poliza, config.email_trust, html=html)


def invitacion(request):
    if request.method == 'POST':
        customers = request.POST.getlist('customer')
        content = request.POST.get('email_content').replace('[[', '{{').replace(']]', '}}')
        template = Template(content)
        for i in range(0, len(customers)):
            secret = secrets.token_urlsafe(8)
            c = Cliente.objects.get(id=customers[i])
            user = c.user
            user.set_password(secret)
            user.save()
            c.cambiar_pass = True
            context = {}
            context['cliente'] = c.to_json()
            context['user'] = {'id': c.user.id, 'username': c.user.username}
            context['password'] = secret
            context = Context(context)
            html = template.render(context)
            send_email('Plan colaborador Banpro/Trust Correduría de Seguros',
                       c.user.email, html)
    return HttpResponseRedirect('/admin/backend/cliente/')


def admin_tasks(request):
    if request.method == "POST":
        if request.POST.get('task') == 'Renovación automática':
            iniciar_proc()
            messages.success(request, "El proceso de renovación automática se ha iniciado con éxito!")
            return JsonResponse({})
        if request.POST.get('task') == 'Notificar pólizas a vencer':
            iniciar_notificacion_vencer()
            messages.success(request, "El proceso de notificacón de pólizas a vencer se ha iniciado con éxito!")
            return JsonResponse({})
