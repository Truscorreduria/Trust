# -*- coding: utf-8 -*-
from django import forms
from openpyxl import Workbook
from django.template.context_processors import csrf
from django.shortcuts import render_to_response
from django.contrib.admin.widgets import AdminDateWidget
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from django.http import HttpResponse, HttpResponseRedirect
from django.db.models import Sum
from backend.models import *
from django.template.defaultfilters import date


class ReporteFecha(forms.Form):
    _referer = forms.CharField(widget=forms.HiddenInput)
    fecha_inicio = forms.DateField(input_formats=['%d/%m/%Y'],
                                   widget=AdminDateWidget(format='%d/%m/%Y'), required=True)
    fecha_fin = forms.DateField(input_formats=['%d/%m/%Y'],
                                widget=AdminDateWidget(format='%d/%m/%Y'), required=True)

    def clean(self):
        if self.cleaned_data['fecha_inicio'] > self.cleaned_data['fecha_fin']:
            raise forms.ValidationError("La fecha inicial no puede ser mayor a la fecha final!")
        else:
            return self.cleaned_data


def render_to_excell(data, filename="Reporte.xlsx"):
    book = Workbook()
    sheet = book.active
    for row in data:
        sheet.append(row)
    response = HttpResponse(content_type="application/ms-excel")
    content = "attachment; filename={0}".format(filename)
    response["Content-Disposition"] = content
    book.save(response)
    return response


def reporte_cotizacion_auto(request):
    message = ""
    form = None

    if 'apply' in request.POST:
        form = ReporteFecha(request.POST)
        if form.is_valid():
            print(form.cleaned_data)
            wb = Workbook()
            ws = wb.active
            c = ws['B1']
            c.font = Font(size=24, underline="single", bold=True)
            c.alignment = Alignment(horizontal='center')
            ws['B1'] = "TRUST CORREDURÍA DE SEGUROS"
            ws.merge_cells('B1:J1')
            ws['B2'] = "Reporte de cotizaciónes"
            ws.merge_cells('B2:J2')
            c = ws['B2']
            c.font = Font(underline="single", bold=True)
            c.alignment = Alignment(horizontal='center')
            ws.sheet_properties.tabColor = "1072BA"
            yellowFill = PatternFill(start_color='ffff00',
                                     end_color='ffff00',
                                     fill_type='solid')
            c.fill = yellowFill
            c = ws['B3']
            c.font = Font(underline="single", bold=True)
            c.alignment = Alignment(horizontal='center')
            ws['B3'] = "Del " + str(form.cleaned_data['fecha_inicio']) + " al " + str(form.cleaned_data['fecha_fin'])
            ws.merge_cells('B3:J3')

            c = ws['A4']
            c.font = Font(underline="single", bold=True)
            c.alignment = Alignment(horizontal='left')
            ws['A4'] = "Cotizador Banpro"
            ws.merge_cells('A4:J4')

            ws['A5'] = 'FECHA'
            ws['B5'] = 'NOMBRE'
            ws['C5'] = 'MARCA'
            ws['D5'] = 'MODELO'
            ws['E5'] = 'AÑO'
            ws['F5'] = 'CHASIS DE REFERENCIA'
            ws['G5'] = 'VALOR DE NUEVO'
            ws['H5'] = 'SUMA ASEGURADA'
            ws['I5'] = 'ROTURA DE VIDRIOS'
            ws['J5'] = 'PRIMA'
            ws['k5'] = 'EMISIÓN'
            ws['L5'] = 'IVA'
            ws['M5'] = 'TOTAL'
            ws['N5'] = 'CUOTA'
            ws['O5'] = 'EXCESO'

            cont = 6

            data = Analitics.objects.filter(created__gte=form.cleaned_data['fecha_inicio'],
                                            created__lte=form.cleaned_data['fecha_fin'])

            for d in data:
                ws.cell(row=cont, column=1).value = str(d.created)
                ws.cell(row=cont, column=2).value = d.user.profile().full_name
                _row = 3
                for r in d.data.split('\n'):
                    ws.cell(row=cont, column=_row).value = r.split(": ")[1]
                    _row += 1
                cont += 1

            nombre_archivo = "Cotizaciones.xlsx"
            response = HttpResponse(content_type="application/ms-excel")
            contenido = "attachment; filename={0}".format(nombre_archivo)
            response["Content-Disposition"] = contenido
            wb.save(response)
            return response
    if not form:
        referer = request.META.get('HTTP_REFERER')
        form = ReporteFecha(initial={'_referer': referer})

    data = {'form': form,
            'header_tittle': 'Por Favor seleccione las fechas en las cual desea generar el reporte'}

    data.update(csrf(request))
    return render_to_response('reports/form.html', data)


def reporte_debito_automatico(request):
    message = ""
    form = None

    if 'apply' in request.POST:
        form = ReporteFecha(request.POST)
        if form.is_valid():

            data = [
                ['numero de orden',
                 'fecha de solicitud',
                 'nombre del asegurado titular de la tarjeta de credito',
                 'numero de identificacion',
                 'numero de tarjeta',
                 'vencimiento de tarjeta',
                 'numero de cuotas',
                 'tipo de tarjeta',
                 'Banco emisor',
                 'moneda',
                 'numero de poliza',
                 'telefono',
                 'email'],
            ]

            inicial = datetime.strptime(request.POST.get('fecha_inicio'), '%d/%m/%Y').strftime('%Y-%m-%d')
            final = (datetime.strptime(request.POST.get('fecha_fin'), '%d/%m/%Y')
                     + timedelta(days=1)).strftime('%Y-%m-%d')

            polizas = Poliza.objects.filter(created__gte=inicial, created__lte=final, medio_pago='debito_automatico')

            for p in polizas:
                data.append([
                    p.print_code(),
                    p.created.strftime('%d/%m/%Y'),
                    p.nombre_asegurado(),
                    p.cedula,
                    p.card_number,
                    p.card_expiry,
                    p.cuotas,
                    p.card_type,
                    p.banco_emisor,
                    p.moneda_cobro,
                    p.no_poliza,
                    p.celular,
                    p.user.email,
                ])
            return render_to_excell(data, 'Debito Automatico.xlsx')
    if not form:
        referer = request.META.get('HTTP_REFERER')
        form = ReporteFecha(initial={'_referer': referer})

    data = {'form': form,
            'header_tittle': 'Por Favor seleccione las fechas en las cual desea generar el reporte'}

    data.update(csrf(request))
    return render_to_response('reports/form.html', data)


def reporte_deduccion_nomina(request):
    message = ""
    form = None

    if 'apply' in request.POST:
        form = ReporteFecha(request.POST)
        if form.is_valid():

            data = [
                ['# EMPL.',
                 'Nombre',
                 'Monto',
                 'Moneda',
                 'Cant. Cuotas a deducir',
                 '# POLIZA',
                 'VIGENCIA'],
            ]

            inicial = datetime.strptime(request.POST.get('fecha_inicio'), '%d/%m/%Y').strftime('%Y-%m-%d')
            final = (datetime.strptime(request.POST.get('fecha_fin'), '%d/%m/%Y')
                     + timedelta(days=1)).strftime('%Y-%m-%d')

            polizas = Poliza.objects.filter(created__gte=inicial, created__lte=final, medio_pago='deduccion_nomina',
                                            cliente__isnull=False)

            for p in polizas:
                data.append([
                    p.fecha_emision,
                    p.cliente.codigo_empleado,
                    p.cliente.full_name,
                    p.monto_cuota,
                    p.moneda_cobro,
                    p.cuotas,
                    p.no_poliza,
                    p.fecha_vence
                ])
            return render_to_excell(data, 'Deducción de nómina.xlsx')
    if not form:
        referer = request.META.get('HTTP_REFERER')
        form = ReporteFecha(initial={'_referer': referer})

    data = {'form': form,
            'header_tittle': 'Por Favor seleccione las fechas en las cual desea generar el reporte'}

    data.update(csrf(request))
    return render_to_response('reports/form.html', data)


def reporte_polizas_vencer(request):
    message = ""
    form = None

    if 'apply' in request.POST:
        form = ReporteFecha(request.POST)
        if form.is_valid():

            data = [
                ['numero de orden',
                 'fecha de solicitud',
                 'nombre del asegurado',
                 'cédula',
                 'numero de poliza',
                 'telefono',
                 'email'],
            ]

            inicial = datetime.strptime(request.POST.get('fecha_inicio'), '%d/%m/%Y').strftime('%Y-%m-%d')
            final = (datetime.strptime(request.POST.get('fecha_fin'), '%d/%m/%Y')
                     + timedelta(days=1)).strftime('%Y-%m-%d')

            polizas = Poliza.objects.filter(fecha_vence__gte=inicial, fecha_vence__lte=final)

            for p in polizas:
                data.append([
                    p.print_code(),
                    p.created.strftime('%d/%m/%Y'),
                    p.nombre_asegurado(),
                    p.cedula,
                    p.no_poliza,
                    p.celular,
                    p.cliente.email_personal,
                ])
            return render_to_excell(data, 'Deducción de nómina.xlsx')
    if not form:
        referer = request.META.get('HTTP_REFERER')
        form = ReporteFecha(initial={'_referer': referer})

    data = {'form': form,
            'header_tittle': 'Por Favor seleccione las fechas en las cual desea generar el reporte'}

    data.update(csrf(request))
    return render_to_response('reports/form.html', data)
